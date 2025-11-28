from fastapi import APIRouter, Request, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import zipstream
import datetime
from io import BytesIO
from openpyxl import Workbook

from app.database import get_db
from app.service.data_service import (
    load_lot_rows_all,
    build_multi_summary,
    build_unit_groups,
    calculate_under_over_kill,
)

router = APIRouter(prefix="/api/pms", tags=["PMS ZIP Export"])


def build_rawdata_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "RAWDATA"

    if not rows:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    sample = rows[0]

    if hasattr(sample, "_mapping"):
        first = sample._mapping
        headers = list(first.keys())
        ws.append(headers)

        for raw in rows:
            ws.append([raw._mapping.get(h) for h in headers])

    elif isinstance(sample, tuple):
        first = sample[0]._mapping
        headers = list(first.keys())
        ws.append(headers)

        for r in rows:
            raw = r[0]
            ws.append([raw._mapping.get(h) for h in headers])

    else:

        headers = list(sample.keys())
        ws.append(headers)
        for raw in rows:
            ws.append([raw.get(h) for h in headers])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_summary_excel(summary_rows, today_str: str) -> StreamingResponse:

    wb = Workbook()
    ws = wb.active
    ws.title = "SUMMARY"

    if summary_rows:
        ws.append(list(summary_rows[0].keys()))
        for r in summary_rows:
            ws.append(list(r.values()))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="SUMMARY_{today_str}.xlsx"'
        },
    )


@router.post("/download-zip")
async def download_zip(request: Request, db: Session = Depends(get_db)):
    payload = await request.json()
    test_list = payload["testList"]
    excel_options = payload["excelOptions"]

    today = datetime.date.today()
    today_str = today.strftime("%Y%m%d")

    summary_rows = []
    for t in test_list:
        test_id = t["id"]
        row = build_multi_summary(db, test_id)
        row["test_id"] = test_id
        summary_rows.append(row)

    summary_only = (
        excel_options.get("summary") is True
        and not excel_options.get("rawdata")
        and not excel_options.get("underkill")
        and not excel_options.get("overkill")
    )

    if summary_only:
        return build_summary_excel(summary_rows, today_str)

    z = zipstream.ZipFile(mode="w", compression=zipstream.ZIP_DEFLATED)

    for t in test_list:
        test_id = t["id"]
        folder = f"{t['lot']}_{t['version']}"

        rows_map = load_lot_rows_all(db, test_id)
        all_rows = rows_map.get(test_id, [])

        if excel_options.get("rawdata") and all_rows:
            raw_excel = build_rawdata_excel(all_rows)
            z.write_iter(f"{folder}/{folder}_rawdata.xlsx", iter([raw_excel]))

        if excel_options.get("underkill") or excel_options.get("overkill"):
            unit_groups, unit_rows = build_unit_groups(all_rows)
            underkill, overkill = calculate_under_over_kill(unit_rows)

            if excel_options.get("underkill") and underkill:
                under_excel = build_rawdata_excel(underkill)
                z.write_iter(f"{folder}/{folder}_underkill.xlsx", iter([under_excel]))

            if excel_options.get("overkill") and overkill:
                over_excel = build_rawdata_excel(overkill)
                z.write_iter(f"{folder}/{folder}_overkill.xlsx", iter([over_excel]))

    if excel_options.get("summary") and summary_rows:
        wb = Workbook()
        ws = wb.active
        ws.title = "SUMMARY"
        ws.append(list(summary_rows[0].keys()))
        for r in summary_rows:
            ws.append(list(r.values()))
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        z.write_iter(f"SUMMARY_{today_str}.xlsx", iter([buf.getvalue()]))

    return StreamingResponse(
        z,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="AI_RESULT_{today_str}.zip"'
        },
    )
